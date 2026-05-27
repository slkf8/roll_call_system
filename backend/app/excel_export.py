from io import BytesIO

from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_to_tuple

from app.schemas import ExcelFillTemplatePayload


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _ensure_writable_merged_cell(worksheet, cell_address: str) -> None:
    row, column = coordinate_to_tuple(cell_address)

    for merged_range in worksheet.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= column <= merged_range.max_col
        ):
            is_top_left = row == merged_range.min_row and column == merged_range.min_col
            if not is_top_left:
                raise HTTPException(
                    status_code=400,
                    detail="Cannot write to a merged cell unless it is the top-left cell",
                )
            return


def fill_excel_template(file_bytes: bytes, payload: ExcelFillTemplatePayload) -> bytes:
    try:
        workbook = load_workbook(BytesIO(file_bytes))
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to read xlsx workbook") from exc

    if payload.worksheetName not in workbook.sheetnames:
        raise HTTPException(status_code=400, detail="Worksheet not found")

    worksheet = workbook[payload.worksheetName]

    for write in payload.writes:
        _ensure_writable_merged_cell(worksheet, write.cellAddress)
        worksheet[write.cellAddress].value = write.value

    output = BytesIO()
    try:
        workbook.save(output)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Unable to write xlsx workbook") from exc

    return output.getvalue()
