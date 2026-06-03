import json
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from app.excel_export import XLSX_MEDIA_TYPE
from app.main import app


client = TestClient(app)


def make_workbook_bytes(*, merged: bool = False) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "原值"
    worksheet["A1"].font = Font(bold=True)
    worksheet["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    worksheet["B2"] = "保留"
    worksheet["C3"] = 9
    worksheet.column_dimensions["A"].width = 24

    if merged:
        worksheet.merge_cells("D4:E4")
        worksheet["D4"] = "合併"

    extra = workbook.create_sheet("Extra")
    extra["A1"] = "extra"

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def base_payload(**overrides):
    payload = {
        "worksheetName": "Sheet1",
        "month": "2026-06",
        "writes": [
            {
                "cellAddress": "C3",
                "value": 4,
                "studentId": 1,
                "studentName": "陳小明",
                "birthday": "2012-03-08",
                "reason": "direct_service_count",
            }
        ],
        "options": {"preserveTemplate": True},
    }
    payload.update(overrides)
    return payload


def post_export(
    *,
    payload=None,
    payload_text: str | None = None,
    file_bytes: bytes | None = None,
    filename: str = "template.xlsx",
):
    data = payload_text if payload_text is not None else json.dumps(payload or base_payload())
    return client.post(
        "/api/exports/excel/fill-template",
        files={
            "file": (
                filename,
                file_bytes if file_bytes is not None else make_workbook_bytes(),
                XLSX_MEDIA_TYPE,
            ),
            "payload": (None, data, "application/json"),
        },
    )


def load_response_workbook(response):
    return load_workbook(BytesIO(response.content))


def test_upload_xlsx_template_success():
    response = post_export()

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["C3"].value == 4


def test_existing_worksheet_can_be_written():
    response = post_export(payload=base_payload(worksheetName="Extra"))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Extra"]["C3"].value == 4


def test_missing_worksheet_returns_400():
    response = post_export(payload=base_payload(worksheetName="Missing"))

    assert response.status_code == 400


def test_writes_single_cell():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": 12}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == 12


def test_writes_multiple_cells():
    response = post_export(
        payload=base_payload(
            writes=[
                {"cellAddress": "A1", "value": 12},
                {"cellAddress": "C3", "value": 5},
            ]
        )
    )

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == 12
    assert workbook["Sheet1"]["C3"].value == 5


def test_empty_writes_returns_400():
    response = post_export(payload=base_payload(writes=[]))

    assert response.status_code == 400


def test_empty_cell_address_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "", "value": 1}]))

    assert response.status_code == 400


def test_invalid_cell_address_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "not-a-cell", "value": 1}]))

    assert response.status_code == 400


def test_cell_range_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1:B2", "value": 1}]))

    assert response.status_code == 400


def test_missing_value_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1"}]))

    assert response.status_code == 400


def test_null_value_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": None}]))

    assert response.status_code == 400


def test_non_number_value_returns_400():
    # Plain "4" is not a valid materials reason string ("<day>-<code>") either.
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "4"}]))

    assert response.status_code == 400


def test_valid_reason_string_can_be_written():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "3-4"}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == "3-4"


def test_multiple_reason_pairs_can_be_written():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "3-4;18-2"}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == "3-4;18-2"


def test_same_day_duplicate_reason_pairs_can_be_written():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "3-4;3-4"}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == "3-4;3-4"


def test_boolean_value_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": True}]))

    assert response.status_code == 400


def test_arbitrary_text_value_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "生病"}]))

    assert response.status_code == 400


def test_empty_string_value_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": ""}]))

    assert response.status_code == 400


def test_formula_like_string_value_returns_400():
    for formula in ("=SUM(A1:A3)", "+1+1", "@formula"):
        response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": formula}]))
        assert response.status_code == 400, formula


def test_reason_code_out_of_range_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "3-7"}]))

    assert response.status_code == 400


def test_trailing_semicolon_reason_string_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "3-4;"}]))

    assert response.status_code == 400


def test_reason_string_over_max_length_returns_400():
    long_value = ";".join(["3-4"] * 300)  # well over 1024 chars
    assert len(long_value) > 1024
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": long_value}]))

    assert response.status_code == 400


def test_zero_day_reason_string_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "0-4"}]))

    assert response.status_code == 400


def test_leading_zero_day_reason_string_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "00-4"}]))

    assert response.status_code == 400


def test_day_above_31_reason_string_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "32-4"}]))

    assert response.status_code == 400


def test_two_digit_day_above_31_reason_string_returns_400():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "A1", "value": "99-6"}]))

    assert response.status_code == 400


def test_preserves_style_when_writing_reason_string():
    response = post_export(
        payload=base_payload(
            writes=[
                {"cellAddress": "C3", "value": 4},
                {"cellAddress": "A1", "value": "3-4;18-2"},
            ]
        )
    )

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == "3-4;18-2"
    assert workbook["Sheet1"]["A1"].font.bold is True
    assert workbook["Sheet1"]["A1"].fill.fgColor.rgb == "00FFFF00"
    assert workbook["Sheet1"]["B2"].value == "保留"
    assert workbook["Sheet1"].column_dimensions["A"].width == 24


def test_preserves_other_cell_values():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "C3", "value": 4}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].value == "原值"
    assert workbook["Sheet1"]["B2"].value == "保留"


def test_preserves_basic_style():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "C3", "value": 4}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["A1"].font.bold is True
    assert workbook["Sheet1"]["A1"].fill.fgColor.rgb == "00FFFF00"


def test_preserves_column_width():
    response = post_export(payload=base_payload(writes=[{"cellAddress": "C3", "value": 4}]))

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"].column_dimensions["A"].width == 24


def test_response_headers_are_xlsx_download():
    response = post_export()

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert response.headers["content-disposition"] == (
        'attachment; filename="filled-template-2026-06.xlsx"'
    )


def test_non_xlsx_file_returns_400():
    response = post_export(file_bytes=b"not excel", filename="template.xls")

    assert response.status_code == 400


def test_malformed_payload_returns_400():
    response = post_export(payload_text="{bad json")

    assert response.status_code == 400


def test_merged_cell_top_left_can_be_written():
    response = post_export(
        file_bytes=make_workbook_bytes(merged=True),
        payload=base_payload(writes=[{"cellAddress": "D4", "value": 8}]),
    )

    assert response.status_code == 200, response.text
    workbook = load_response_workbook(response)
    assert workbook["Sheet1"]["D4"].value == 8


def test_merged_cell_non_top_left_returns_400():
    response = post_export(
        file_bytes=make_workbook_bytes(merged=True),
        payload=base_payload(writes=[{"cellAddress": "E4", "value": 8}]),
    )

    assert response.status_code == 400
